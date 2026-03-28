import csv
import io
import re
from collections import Counter
from datetime import datetime

import frappe
from frappe import _
from frappe.utils import add_days, cstr, get_datetime, getdate


REQUIRED_FIELD_MAP = {
	"biometric_employee_code": [
		"biometric_employee_code",
		"employee_code",
		"emp_code",
		"employee_id",
		"biometric_code",
		"enroll_no",
		"enroll_id",
	],
	"punch_timestamp": [
		"punch_timestamp",
		"timestamp",
		"datetime",
		"punch_datetime",
		"log_datetime",
	],
	"punch_date": ["punch_date", "date", "log_date", "attendance_date"],
	"punch_time": ["punch_time", "time", "log_time"],
	"direction": ["direction", "log_type", "in_out", "punch_direction"],
	"device_id": ["device_id", "terminal_id", "machine_id"],
	"device_location": ["device_location", "location", "device_name", "terminal_name"],
}

ESSENTIAL_FIELDS = ("biometric_employee_code",)


class AttendanceImportPreviewService:
	def get_dashboard_context(self):
		settings = self.get_settings()
		last_successful_batch = frappe.get_all(
			"Attendance Import Batch",
			filters={"processing_status": "Processed"},
			fields=["name", "source_file_name", "file_date_from", "file_date_to", "modified"],
			order_by="modified desc",
			limit=1,
		)

		return {
			"last_uploaded_till_date": settings.get("last_attendance_uploaded_till_date"),
			"pending_from_date": settings.get("next_pending_attendance_from_date"),
			"weekly_upload_day": settings.get("weekly_upload_day"),
			"pending_upload_alert_days": settings.get("pending_upload_alert_days"),
			"last_successful_batch": last_successful_batch[0] if last_successful_batch else None,
			"unresolved_corrections": frappe.db.count(
				"Attendance Correction Request",
				{"approval_status": ["in", ["Draft", "Submitted"]]},
			),
			"unprocessed_batches": frappe.db.count(
				"Attendance Import Batch",
				{"processing_status": ["in", ["Draft", "Validated", "Partially Processed"]]},
			),
		}

	def preview_file(self, file_url: str):
		file_doc = self.get_file_doc(file_url)
		rows, source_file_name = self.read_rows(file_doc)
		settings = self.get_settings()

		preview = {
			"source_file_name": source_file_name,
			"total_rows": len(rows),
			"imported_rows": 0,
			"rejected_rows": 0,
			"duplicate_rows": 0,
			"unmatched_employee_rows": 0,
			"invalid_timestamp_rows": 0,
			"future_date_rows": 0,
			"overlap_rows": 0,
			"locked_period_rows": 0,
			"file_date_from": None,
			"file_date_to": None,
			"last_uploaded_till_date": settings.get("last_attendance_uploaded_till_date"),
			"pending_from_date": settings.get("next_pending_attendance_from_date"),
			"suggested_upload_to_date": None,
			"columns_found": [],
			"required_columns_missing": [],
			"fatal_errors": [],
			"warnings": [],
			"unmatched_employee_codes": [],
			"duplicate_samples": [],
			"invalid_timestamp_samples": [],
			"future_date_samples": [],
			"overlap_samples": [],
			"locked_period_samples": [],
			"sample_rows": [],
		}

		if not rows:
			preview["fatal_errors"].append(_("The uploaded file does not contain any rows."))
			return preview

		header_map = self.map_headers(rows[0].keys())
		preview["columns_found"] = list(rows[0].keys())
		preview["required_columns_missing"] = self.get_missing_required_columns(header_map)

		if preview["required_columns_missing"]:
			preview["fatal_errors"].append(
				_("Missing required columns: {0}").format(", ".join(preview["required_columns_missing"]))
			)
			return preview

		if not self.employee_code_field_available():
			preview["warnings"].append(
				_(
					"Employee custom field custom_biometric_employee_code was not found. "
					"Employee matching will not work until the custom field is created."
				)
			)

		seen_timestamps = Counter()
		employee_codes = set()
		normalized_rows = []

		for idx, raw_row in enumerate(rows, start=2):
			normalized = self.normalize_row(raw_row, header_map, idx)
			normalized_rows.append(normalized)
			if normalized["biometric_employee_code"]:
				employee_codes.add(normalized["biometric_employee_code"])
			if normalized["timestamp_key"]:
				seen_timestamps[normalized["timestamp_key"]] += 1

		known_codes = self.get_known_employee_codes(employee_codes)
		locked_dates = self.get_locked_attendance_dates()
		last_uploaded_till = settings.get("last_attendance_uploaded_till_date")

		for normalized in normalized_rows:
			if len(preview["sample_rows"]) < 5:
				preview["sample_rows"].append(
					{
						"row_number": normalized["row_number"],
						"biometric_employee_code": normalized["biometric_employee_code"],
						"punch_timestamp": normalized["punch_timestamp"],
						"direction": normalized["direction"],
						"device_id": normalized["device_id"],
					}
				)

			if normalized["error"]:
				preview["invalid_timestamp_rows"] += 1
				if len(preview["invalid_timestamp_samples"]) < 10:
					preview["invalid_timestamp_samples"].append(
						self._sample(normalized, normalized["error"])
					)
				continue

			row_date = normalized["punch_datetime"].date()
			preview["file_date_from"] = (
				row_date
				if not preview["file_date_from"]
				else min(preview["file_date_from"], row_date)
			)
			preview["file_date_to"] = (
				row_date
				if not preview["file_date_to"]
				else max(preview["file_date_to"], row_date)
			)

			if normalized["timestamp_key"] and seen_timestamps[normalized["timestamp_key"]] > 1:
				preview["duplicate_rows"] += 1
				if len(preview["duplicate_samples"]) < 10:
					preview["duplicate_samples"].append(
						self._sample(normalized, _("Duplicate employee and timestamp in file"))
					)

			if normalized["biometric_employee_code"] not in known_codes:
				preview["unmatched_employee_rows"] += 1
				if (
					normalized["biometric_employee_code"]
					and normalized["biometric_employee_code"] not in preview["unmatched_employee_codes"]
				):
					preview["unmatched_employee_codes"].append(normalized["biometric_employee_code"])

			if row_date > getdate():
				preview["future_date_rows"] += 1
				if len(preview["future_date_samples"]) < 10:
					preview["future_date_samples"].append(
						self._sample(normalized, _("Future date rows are not allowed"))
					)

			if last_uploaded_till and row_date <= last_uploaded_till:
				preview["overlap_rows"] += 1
				if len(preview["overlap_samples"]) < 10:
					preview["overlap_samples"].append(
						self._sample(
							normalized,
							_("Row overlaps with already covered attendance period"),
						)
					)

			if row_date in locked_dates:
				preview["locked_period_rows"] += 1
				if len(preview["locked_period_samples"]) < 10:
					preview["locked_period_samples"].append(
						self._sample(normalized, _("Attendance is locked for this date"))
					)

		preview["rejected_rows"] = (
			preview["invalid_timestamp_rows"]
			+ preview["future_date_rows"]
			+ preview["locked_period_rows"]
		)
		preview["imported_rows"] = max(preview["total_rows"] - preview["rejected_rows"], 0)
		if preview["file_date_to"]:
			preview["suggested_upload_to_date"] = preview["file_date_to"]

		if preview["duplicate_rows"]:
			preview["warnings"].append(
				_("{0} duplicate row(s) were found in the uploaded file.").format(preview["duplicate_rows"])
			)
		if preview["overlap_rows"]:
			preview["warnings"].append(
				_("{0} row(s) fall in an already covered attendance period.").format(preview["overlap_rows"])
			)
		if preview["unmatched_employee_rows"]:
			preview["warnings"].append(
				_("{0} row(s) could not be matched to an employee.").format(
					preview["unmatched_employee_rows"]
				)
			)

		return preview

	def build_error_summary(self, preview):
		lines = []
		if preview.get("fatal_errors"):
			lines.extend(preview["fatal_errors"])
		if preview.get("warnings"):
			lines.extend(preview["warnings"])
		return "\n".join(lines)

	def get_settings(self):
		defaults = {
			"last_attendance_uploaded_till_date": None,
			"next_pending_attendance_from_date": None,
			"weekly_upload_day": "Saturday",
			"pending_upload_alert_days": 7,
		}

		if not frappe.db.exists("DocType", "Attendance Control Settings"):
			return defaults

		try:
			doc = frappe.get_single("Attendance Control Settings")
		except Exception:
			return defaults

		defaults.update(
			{
				"last_attendance_uploaded_till_date": doc.get("last_attendance_uploaded_till_date"),
				"next_pending_attendance_from_date": doc.get("next_pending_attendance_from_date"),
				"weekly_upload_day": doc.get("weekly_upload_day") or "Saturday",
				"pending_upload_alert_days": doc.get("pending_upload_alert_days") or 7,
			}
		)

		if (
			not defaults["next_pending_attendance_from_date"]
			and defaults["last_attendance_uploaded_till_date"]
		):
			defaults["next_pending_attendance_from_date"] = add_days(
				defaults["last_attendance_uploaded_till_date"], 1
			)

		return defaults

	def get_file_doc(self, file_url):
		if frappe.db.exists("File", {"file_url": file_url}):
			return frappe.get_doc("File", {"file_url": file_url})
		if frappe.db.exists("File", file_url):
			return frappe.get_doc("File", file_url)
		frappe.throw(_("Uploaded file could not be found."))

	def read_rows(self, file_doc):
		file_name = file_doc.file_name or file_doc.file_url
		extension = cstr(file_name).lower().rsplit(".", 1)[-1] if "." in cstr(file_name) else ""
		content = file_doc.get_content()

		if extension == "csv":
			return self.read_csv_rows(content), file_name
		if extension == "xlsx":
			return self.read_xlsx_rows(content), file_name

		frappe.throw(_("Only CSV and XLSX files are supported for preview right now."))

	def read_csv_rows(self, content):
		if isinstance(content, bytes):
			text = content.decode("utf-8-sig")
		else:
			text = cstr(content)
		reader = csv.DictReader(io.StringIO(text))
		return [row for row in reader if any(cstr(value).strip() for value in row.values())]

	def read_xlsx_rows(self, content):
		try:
			from openpyxl import load_workbook
		except ImportError:
			frappe.throw(_("openpyxl is required to preview XLSX attendance files."))

		workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
		sheet = workbook.active
		rows = list(sheet.iter_rows(values_only=True))
		if not rows:
			return []

		headers = [cstr(value).strip() for value in rows[0]]
		data = []
		for row in rows[1:]:
			record = {}
			for idx, header in enumerate(headers):
				if not header:
					continue
				record[header] = row[idx] if idx < len(row) else None
			if any(cstr(value).strip() for value in record.values()):
				data.append(record)
		return data

	def map_headers(self, headers):
		normalized = {self.normalize_header(header): header for header in headers}
		mapped = {}
		for target, aliases in REQUIRED_FIELD_MAP.items():
			for alias in aliases:
				if alias in normalized:
					mapped[target] = normalized[alias]
					break
		return mapped

	def get_missing_required_columns(self, header_map):
		missing = []
		for fieldname in ESSENTIAL_FIELDS:
			if fieldname not in header_map:
				missing.append(fieldname)

		has_timestamp = "punch_timestamp" in header_map
		has_date_and_time = "punch_date" in header_map and "punch_time" in header_map
		if not has_timestamp and not has_date_and_time:
			missing.extend(["punch_timestamp or punch_date + punch_time"])

		return missing

	def normalize_row(self, raw_row, header_map, row_number):
		row = {self.normalize_header(key): value for key, value in raw_row.items()}

		employee_code = self.clean_value(row.get(self.normalize_header(header_map.get("biometric_employee_code"))))
		punch_timestamp = self.parse_timestamp(row, header_map)
		direction = self.clean_value(row.get(self.normalize_header(header_map.get("direction"))))
		device_id = self.clean_value(row.get(self.normalize_header(header_map.get("device_id"))))
		device_location = self.clean_value(
			row.get(self.normalize_header(header_map.get("device_location")))
		)

		error = None
		if not employee_code:
			error = _("Biometric employee code is missing")
		elif not punch_timestamp:
			error = _("Punch timestamp could not be parsed")

		return {
			"row_number": row_number,
			"biometric_employee_code": employee_code,
			"punch_datetime": punch_timestamp,
			"punch_timestamp": punch_timestamp.isoformat(sep=" ") if punch_timestamp else None,
			"direction": direction,
			"device_id": device_id,
			"device_location": device_location,
			"timestamp_key": f"{employee_code}::{punch_timestamp.isoformat()}" if employee_code and punch_timestamp else None,
			"error": error,
		}

	def parse_timestamp(self, row, header_map):
		timestamp_header = header_map.get("punch_timestamp")
		date_header = header_map.get("punch_date")
		time_header = header_map.get("punch_time")

		if timestamp_header:
			value = row.get(self.normalize_header(timestamp_header))
			return self._coerce_datetime(value)

		date_value = row.get(self.normalize_header(date_header)) if date_header else None
		time_value = row.get(self.normalize_header(time_header)) if time_header else None
		if not date_value or not time_value:
			return None

		return self._coerce_datetime(f"{date_value} {time_value}")

	def _coerce_datetime(self, value):
		if not value:
			return None
		if isinstance(value, datetime):
			return value
		try:
			return get_datetime(value)
		except Exception:
			return None

	def employee_code_field_available(self):
		return frappe.db.has_column("Employee", "custom_biometric_employee_code")

	def get_known_employee_codes(self, employee_codes):
		if not employee_codes or not self.employee_code_field_available():
			return set()

		return set(
			frappe.get_all(
				"Employee",
				filters={"custom_biometric_employee_code": ["in", list(employee_codes)]},
				pluck="custom_biometric_employee_code",
			)
		)

	def get_locked_attendance_dates(self):
		if not frappe.db.exists("DocType", "Attendance"):
			return set()
		if not frappe.db.has_column("Attendance", "custom_attendance_locked"):
			return set()

		return set(
			frappe.get_all(
				"Attendance",
				filters={"custom_attendance_locked": 1},
				pluck="attendance_date",
			)
		)

	def clean_value(self, value):
		return cstr(value).strip() if value is not None else ""

	def normalize_header(self, header):
		header = cstr(header).strip().lower()
		header = re.sub(r"[^a-z0-9]+", "_", header)
		return header.strip("_")

	def _sample(self, normalized, message):
		return {
			"row_number": normalized["row_number"],
			"employee_code": normalized["biometric_employee_code"],
			"punch_timestamp": normalized["punch_timestamp"],
			"message": message,
		}
